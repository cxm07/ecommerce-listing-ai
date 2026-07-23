"""V1 application services, in-memory persistence, and Excel adapters."""
from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
from threading import RLock
from typing import Any, Protocol
from copy import deepcopy
from uuid import UUID, uuid4

from openpyxl import Workbook, load_workbook

from app.workflow import TaskStatus, WorkflowService


def now() -> str:
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def clean(value: Any) -> str | None:
    if value is None:
        return None
    result = " ".join(str(value).split())
    return result or None


def json_value(value: Any) -> Any:
    if isinstance(value, UUID): return str(value)
    if isinstance(value, Decimal): return float(value.quantize(Decimal("0.01")))
    if isinstance(value, list): return [json_value(item) for item in value]
    if isinstance(value, dict): return {key: json_value(item) for key, item in value.items()}
    return value


class DomainError(Exception):
    def __init__(self, code: str, message: str, status: int = 400, details: dict[str, Any] | None = None):
        self.code, self.message, self.status, self.details = code, message, status, details
        super().__init__(message)


@dataclass
class Task:
    id: UUID; task_name: str; category: str; creator_id: str; status: TaskStatus = TaskStatus.DRAFT
    created_at: str = field(default_factory=now); updated_at: str = field(default_factory=now)
    version: int = 1

@dataclass
class TaskFile:
    id: UUID; task_id: UUID; storage_path: str; original_filename: str; file_kind: str; created_at: str = field(default_factory=now)

@dataclass
class Product:
    id: UUID; task_id: UUID; product_name: str | None; category: str | None; material: str | None
    source_row: int; source_payload: dict[str, Any]; created_at: str = field(default_factory=now); updated_at: str = field(default_factory=now)

@dataclass
class SKU:
    id: UUID; product_id: UUID; sku_code: str | None; color: str | None; size: str | None
    price: Decimal | None; stock: int | None; source_row: int; source_payload: dict[str, Any]
    created_at: str = field(default_factory=now); updated_at: str = field(default_factory=now)

@dataclass
class Issue:
    id: UUID; task_id: UUID; product_id: UUID | None; sku_id: UUID | None; code: str; field: str
    severity: str; message: str; source_ref: dict[str, Any]; signature: str; resolved: bool = False; created_at: str = field(default_factory=now)

@dataclass
class GeneratedContent:
    id: UUID; task_id: UUID; product_id: UUID; title: str; selling_points: list[str]; unsupported_claims: list[str]
    model_metadata: dict[str, Any]; created_at: str = field(default_factory=now)

@dataclass
class Approval:
    id: UUID; task_id: UUID; reviewer_id: str; approval_type: str; decision: str; comment: str | None; created_at: str = field(default_factory=now)

@dataclass
class AuditLog:
    id: UUID; task_id: UUID; actor_id: str | None; action: str; source_ref: dict[str, Any] | None; created_at: str = field(default_factory=now)


class Repository(Protocol):
    def get_task(self, task_id: UUID) -> Task: ...
    def list_tasks(self) -> list[Task]: ...
    def get_product(self, product_id: UUID) -> Product: ...
    def get_sku(self, sku_id: UUID) -> SKU: ...
    def unit_of_work(self) -> "UnitOfWork": ...


class UnitOfWork(Protocol):
    def __enter__(self) -> "UnitOfWork": ...
    def __exit__(self, exc_type: object, exc: object, traceback: object) -> bool: ...
    def commit(self) -> None: ...


class MemoryUnitOfWork:
    def __init__(self, repository: "MemoryRepository") -> None:
        self.repository, self.committed = repository, False
        self.snapshot: tuple[dict[UUID, Any], ...] | None = None

    def __enter__(self) -> "MemoryUnitOfWork":
        with self.repository._lock:
            self.snapshot = deepcopy((self.repository._tasks, self.repository._files, self.repository._products, self.repository._skus, self.repository._issues, self.repository._contents, self.repository._approvals, self.repository._audits))
        return self

    def commit(self) -> None:
        with self.repository._lock:
            self.committed = True

    def __exit__(self, exc_type: object, exc: object, traceback: object) -> bool:
        if exc_type is not None or not self.committed:
            assert self.snapshot is not None
            with self.repository._lock:
                (self.repository._tasks, self.repository._files, self.repository._products, self.repository._skus, self.repository._issues, self.repository._contents, self.repository._approvals, self.repository._audits) = self.snapshot
        return False


class MemoryRepository:
    def __init__(self) -> None:
        self._tasks: dict[UUID, Task] = {}; self._files: dict[UUID, TaskFile] = {}; self._products: dict[UUID, Product] = {}
        self._skus: dict[UUID, SKU] = {}; self._issues: dict[UUID, Issue] = {}; self._contents: dict[UUID, GeneratedContent] = {}
        self._approvals: dict[UUID, Approval] = {}; self._audits: dict[UUID, AuditLog] = {}; self._lock = RLock()

    def get_task(self, task_id: UUID) -> Task:
        try: return self._tasks[task_id]
        except KeyError: raise DomainError("TASK_NOT_FOUND", "未找到任务", 404)
    def get_product(self, product_id: UUID) -> Product:
        try: return self._products[product_id]
        except KeyError: raise DomainError("PRODUCT_NOT_FOUND", "未找到商品", 404)
    def get_sku(self, sku_id: UUID) -> SKU:
        try: return self._skus[sku_id]
        except KeyError: raise DomainError("SKU_NOT_FOUND", "未找到 SKU", 404)
    def list_tasks(self) -> list[Task]: return list(self._tasks.values())
    def list_task_files(self, task_id: UUID) -> list[TaskFile]: return [item for item in self._files.values() if item.task_id == task_id]
    def list_products(self, task_id: UUID) -> list[Product]: return [item for item in self._products.values() if item.task_id == task_id]
    def list_skus(self, task_id: UUID) -> list[SKU]: return [item for item in self._skus.values() if self.get_product(item.product_id).task_id == task_id]
    def list_issues(self, task_id: UUID) -> list[Issue]: return [item for item in self._issues.values() if item.task_id == task_id]
    def list_generated_content(self, task_id: UUID) -> list[GeneratedContent]: return [item for item in self._contents.values() if item.task_id == task_id]
    def list_approvals(self, task_id: UUID) -> list[Approval]: return [item for item in self._approvals.values() if item.task_id == task_id]
    def list_audit_logs(self, task_id: UUID) -> list[AuditLog]: return [item for item in self._audits.values() if item.task_id == task_id]
    def find_issue(self, signature: str) -> Issue | None: return next((item for item in self._issues.values() if item.signature == signature), None)
    def add_task(self, item: Task) -> None: self._tasks[item.id] = item
    def add_file(self, item: TaskFile) -> None: self._files[item.id] = item
    def add_product(self, item: Product) -> None: self._products[item.id] = item
    def add_sku(self, item: SKU) -> None: self._skus[item.id] = item
    def add_issue(self, item: Issue) -> None: self._issues[item.id] = item
    def add_content(self, item: GeneratedContent) -> None: self._contents[item.id] = item
    def add_approval(self, item: Approval) -> None: self._approvals[item.id] = item
    def add_audit(self, item: AuditLog) -> None: self._audits[item.id] = item
    def update_task(self, item: Task) -> None: self._tasks[item.id] = item
    def update_product(self, item: Product) -> None: self._products[item.id] = item
    def update_sku(self, item: SKU) -> None: self._skus[item.id] = item
    def update_issue(self, item: Issue) -> None: self._issues[item.id] = item
    def advance_task(self, task_id: UUID, expected_version: int, new_status: TaskStatus | None = None) -> int:
        task = self.get_task(task_id)
        if task.version != expected_version: raise DomainError("CONCURRENT_MODIFICATION", "任务已被其他请求修改", 409)
        if new_status is not None: task.status = new_status
        task.version += 1; task.updated_at = now(); return task.version
    def unit_of_work(self) -> MemoryUnitOfWork: return MemoryUnitOfWork(self)


class LocalFileStorage:
    def __init__(self, root: Path) -> None: self.root = root
    def save(self, kind: str, payload: bytes) -> str:
        if not payload: raise DomainError("EMPTY_FILE", "上传文件不能为空")
        folder = self.root / ("sources" if kind == "source" else "exports"); folder.mkdir(parents=True, exist_ok=True)
        path = folder / f"{uuid4()}.xlsx"
        path.write_bytes(payload)
        return str(path.relative_to(self.root))
    def read(self, relative: str) -> bytes:
        path = (self.root / relative).resolve()
        if self.root.resolve() not in path.parents or not path.is_file(): raise DomainError("FILE_NOT_FOUND", "文件不存在", 404)
        return path.read_bytes()


def atomic(operation: Any) -> Any:
    """Run one application command against a rollback-capable repository UoW."""
    def wrapped(self: "WorkflowApplication", *args: Any, **kwargs: Any) -> Any:
        with self.repo.unit_of_work() as uow:
            result = operation(self, *args, **kwargs)
            uow.commit()
            return result
    return wrapped


HEADERS = ["product_name", "category", "material", "sku_code", "color", "size", "price", "stock"]

class ExcelSourceAdapter:
    template = "mvp-products-v1"
    def parse(self, source: bytes) -> list[dict[str, Any]]:
        try: book = load_workbook(BytesIO(source), data_only=False)
        except Exception as exc: raise DomainError("INVALID_EXCEL", "无法打开 Excel 文件") from exc
        if "Products" not in book.sheetnames: raise DomainError("INVALID_TEMPLATE", "缺少 Products 工作表")
        sheet = book["Products"]
        headers = [cell.value for cell in sheet[1]]
        if headers != HEADERS: raise DomainError("INVALID_TEMPLATE", "Excel 表头不符合 mvp-products-v1")
        rows: list[dict[str, Any]] = []
        for row_index, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(value is None or str(value).strip() == "" for value in values): continue
            payload = dict(zip(HEADERS, values, strict=True)); rows.append({"row": row_index, "payload": payload})
        return rows


class ExcelOutputAdapter:
    def export(self, workspace: dict[str, Any]) -> bytes:
        book = Workbook(); products = book.active; products.title = "products"
        products.append(["product_id", "product_name", "category", "material", "source_row"])
        for product in workspace["products"]: products.append([product["id"], product["product_name"], product["category"], product["material"], product["source_row"]])
        skus = book.create_sheet("skus"); skus.append(["sku_id", "product_id", "sku_code", "color", "size", "price", "stock", "source_row"])
        for sku in workspace["skus"]: skus.append([sku[k] for k in ["id", "product_id", "sku_code", "color", "size", "price", "stock", "source_row"]])
        copy = book.create_sheet("listing-copy"); copy.append(["product_id", "title", "selling_points", "unsupported_claims"])
        for item in workspace["generated_content"]: copy.append([item["product_id"], item["title"], " | ".join(item["selling_points"]), " | ".join(item["unsupported_claims"])])
        issues = book.create_sheet("issues"); issues.append(["code", "severity", "field", "resolved", "source_ref"])
        for issue in workspace["issues"]: issues.append([issue[k] if k != "source_ref" else str(issue[k]) for k in ["code", "severity", "field", "resolved", "source_ref"]])
        audit = book.create_sheet("audit-summary"); audit.append(["action", "actor_id", "created_at", "source_ref"])
        for item in workspace["audit_logs"]: audit.append([item[k] if k != "source_ref" else str(item[k]) for k in ["action", "actor_id", "created_at", "source_ref"]])
        output = BytesIO(); book.save(output); return output.getvalue()


class WorkflowApplication:
    def __init__(self, repo: MemoryRepository, storage: LocalFileStorage, actor_id: str, max_upload_bytes: int):
        self.repo, self.storage, self.actor_id, self.max_upload_bytes = repo, storage, actor_id, max_upload_bytes
        self.workflow, self.reader, self.writer = WorkflowService(), ExcelSourceAdapter(), ExcelOutputAdapter()
    def audit(self, task_id: UUID, action: str, ref: dict[str, Any] | None = None) -> None:
        log = AuditLog(uuid4(), task_id, self.actor_id, action, ref); self.repo.add_audit(log)
    def list_tasks(self) -> list[Task]: return sorted(self.repo.list_tasks(), key=lambda item: item.updated_at, reverse=True)
    def get_task(self, task_id: UUID) -> Task: return self.repo.get_task(task_id)
    def transition(self, task: Task, target: TaskStatus, action: str) -> None:
        task.status = self.workflow.transition(task.status, target); task.updated_at = now(); self.repo.update_task(task); self.audit(task.id, action)
    @atomic
    def create_task(self, name: str, category: str) -> Task:
        name, category = clean(name), clean(category)
        if not name or not category: raise DomainError("INVALID_TASK", "任务名称和类目不能为空")
        task = Task(uuid4(), name, category, self.actor_id); self.repo.add_task(task); self.audit(task.id, "task_created"); return task
    @atomic
    def upload(self, task_id: UUID, filename: str, payload: bytes) -> TaskFile:
        task = self.repo.get_task(task_id)
        if task.status != TaskStatus.DRAFT: raise DomainError("INVALID_TASK_STATE", "当前任务状态不能上传文件", 409)
        if not filename.lower().endswith(".xlsx"): raise DomainError("INVALID_FILE_TYPE", "仅支持 .xlsx 文件")
        if len(payload) > self.max_upload_bytes: raise DomainError("FILE_TOO_LARGE", "文件超过大小限制")
        if self.repo.list_task_files(task_id): raise DomainError("SOURCE_ALREADY_EXISTS", "任务已有原始文件，不能覆盖", 409)
        item = TaskFile(uuid4(), task_id, self.storage.save("source", payload), filename, "source"); self.repo.add_file(item)
        self.transition(task, TaskStatus.UPLOADED, "source_uploaded"); return item
    @atomic
    def parse(self, task_id: UUID) -> dict[str, int]:
        task = self.repo.get_task(task_id)
        if task.status != TaskStatus.UPLOADED: raise DomainError("INVALID_TASK_STATE", "当前任务状态不能解析", 409)
        source = next((item for item in self.repo.list_task_files(task_id) if item.file_kind == "source"), None)
        if not source: raise DomainError("SOURCE_NOT_FOUND", "未找到原始文件", 404)
        self.transition(task, TaskStatus.PARSING, "parsing_started")
        try:
            rows = self.reader.parse(self.storage.read(source.storage_path)); groups: dict[tuple[str | None, str | None, str | None], Product] = {}
            for record in rows:
                raw = record["payload"]; key = (clean(raw["product_name"]), clean(raw["category"]), clean(raw["material"]))
                product = groups.get(key)
                if product is None:
                    product = Product(uuid4(), task_id, *key, record["row"], raw); groups[key] = product; self.repo.add_product(product)
                try: price = Decimal(str(raw["price"])) if raw["price"] not in (None, "") else None
                except (InvalidOperation, ValueError): price = None
                stock = raw["stock"] if isinstance(raw["stock"], int) and not isinstance(raw["stock"], bool) else None
                sku = SKU(uuid4(), product.id, clean(raw["sku_code"]), clean(raw["color"]), clean(raw["size"]), price, stock, record["row"], raw); self.repo.add_sku(sku)
            self.revalidate(task_id, source); self.transition(task, TaskStatus.WAITING_PRODUCT_REVIEW, "parsing_completed")
        except DomainError:
            self.transition(task, TaskStatus.FAILED, "parsing_failed"); raise
        issues = self.repo.list_issues(task_id); counts = defaultdict(int)
        for issue in issues:
            if not issue.resolved: counts[issue.severity] += 1
        return {"product_count": len(self.repo.list_products(task_id)), "sku_count": len(self.repo.list_skus(task_id)), "issue_count": sum(counts.values()), "error_count": counts["error"], "warning_count": counts["warning"], "info_count": counts["info"]}
    def product_task(self, product_id: UUID) -> UUID: return self.repo.get_product(product_id).task_id
    def revalidate(self, task_id: UUID, source: TaskFile | None = None) -> None:
        source = source or next(item for item in self.repo.list_task_files(task_id) if item.file_kind == "source")
        products = self.repo.list_products(task_id); skus = self.repo.list_skus(task_id)
        wanted: list[tuple[UUID, UUID, str, str, str, str, int]] = []
        counts = defaultdict(int)
        for sku in skus:
            if sku.sku_code: counts[sku.sku_code] += 1
        seen_codes: set[str] = set()
        for sku in skus:
            product = self.repo.get_product(sku.product_id)
            if sku.sku_code and counts[sku.sku_code] > 1 and sku.sku_code in seen_codes: wanted.append((product.id, sku.id, "DUPLICATE_SKU", "sku_code", "error", f"SKU 编码 {sku.sku_code} 重复", sku.source_row))
            if sku.sku_code: seen_codes.add(sku.sku_code)
            if not sku.color: wanted.append((product.id, sku.id, "MISSING_COLOR", "color", "warning", "颜色缺失，需要人工确认", sku.source_row))
            if sku.source_payload.get("price") not in (None, "") and sku.price is None: wanted.append((product.id, sku.id, "INVALID_PRICE", "price", "error", "价格格式无效", sku.source_row))
            if sku.stock is None: wanted.append((product.id, sku.id, "MISSING_STOCK", "stock", "warning", "库存缺失，需要人工确认", sku.source_row))
        for sku in skus:
            product = self.repo.get_product(sku.product_id)
            raw_name = sku.source_payload.get("product_name")
            if clean(raw_name) != raw_name:
                wanted.append((product.id, sku.id, "NORMALIZATION_NEEDED", "product_name", "info", "商品名称已进行空白标准化", sku.source_row))
        signatures = set()
        for product_id, sku_id, code, field_name, severity, message, row in wanted:
            signature = f"{code}:{product_id}:{sku_id if sku_id.int else ''}:{field_name}:{row}"; signatures.add(signature)
            existing = self.repo.find_issue(signature)
            if existing: existing.resolved = False; self.repo.update_issue(existing); continue
            ref = {"file_id": str(source.id), "file_name": source.original_filename, "template": "mvp-products-v1", "sheet": "Products", "row": row, "field": field_name}
            issue = Issue(uuid4(), task_id, product_id, sku_id if sku_id.int else None, code, field_name, severity, message, ref, signature); self.repo.add_issue(issue)
        for issue in self.repo.list_issues(task_id):
            if issue.signature not in signatures: issue.resolved = True; self.repo.update_issue(issue)
    def workspace(self, task_id: UUID) -> dict[str, Any]:
        task = self.repo.get_task(task_id); products = self.repo.list_products(task_id); skus = self.repo.list_skus(task_id)
        return json_value({"task": asdict(task), "files": [asdict(x) for x in self.repo.list_task_files(task_id)], "products": [asdict(x) for x in products], "skus": [asdict(x) for x in skus], "issues": [asdict(x) for x in self.repo.list_issues(task_id)], "generated_content": [asdict(x) for x in self.repo.list_generated_content(task_id)], "approvals": [asdict(x) for x in self.repo.list_approvals(task_id)], "audit_logs": [asdict(x) for x in sorted(self.repo.list_audit_logs(task_id), key=lambda x: x.created_at, reverse=True)]})
    @atomic
    def patch_product(self, product_id: UUID, changes: dict[str, Any]) -> dict[str, Any]:
        product = self.repo.get_product(product_id); task = self.repo.get_task(product.task_id)
        if task.status != TaskStatus.WAITING_PRODUCT_REVIEW: raise DomainError("INVALID_TASK_STATE", "当前状态不能修改商品", 409)
        for key, value in changes.items(): setattr(product, key, clean(value) if isinstance(value, str) else value)
        product.updated_at = now(); task.updated_at = now(); self.repo.update_product(product); self.repo.update_task(task); self.revalidate(task.id); self.audit(task.id, "product_updated", {"product_id": str(product_id)}); return self.workspace(task.id)
    @atomic
    def patch_sku(self, sku_id: UUID, changes: dict[str, Any]) -> dict[str, Any]:
        sku = self.repo.get_sku(sku_id); task = self.repo.get_task(self.product_task(sku.product_id))
        if task.status != TaskStatus.WAITING_PRODUCT_REVIEW: raise DomainError("INVALID_TASK_STATE", "当前状态不能修改 SKU", 409)
        for key, value in changes.items(): setattr(sku, key, clean(value) if isinstance(value, str) else value)
        sku.updated_at = now(); task.updated_at = now(); self.repo.update_sku(sku); self.repo.update_task(task); self.revalidate(task.id); self.audit(task.id, "sku_updated", {"sku_id": str(sku_id)}); return self.workspace(task.id)
    @atomic
    def approve_products(self, task_id: UUID, comment: str | None) -> dict[str, Any]:
        task = self.repo.get_task(task_id)
        if task.status != TaskStatus.WAITING_PRODUCT_REVIEW: raise DomainError("INVALID_TASK_STATE", "当前状态不能审核商品", 409)
        blocking = [asdict(i) for i in self.repo.list_issues(task_id) if not i.resolved and i.severity == "error"]
        if blocking: raise DomainError("UNRESOLVED_ERROR_ISSUES", "仍有错误级问题需要处理", 409, {"issues": json_value(blocking)})
        approval = Approval(uuid4(), task_id, self.actor_id, "product", "approved", comment); self.repo.add_approval(approval); self.transition(task, TaskStatus.PRODUCT_APPROVED, "products_approved"); return self.workspace(task_id)
    @atomic
    def generate_copy(self, task_id: UUID) -> dict[str, Any]:
        task = self.repo.get_task(task_id)
        if task.status != TaskStatus.PRODUCT_APPROVED: raise DomainError("INVALID_TASK_STATE", "请先审核商品", 409)
        self.transition(task, TaskStatus.GENERATING_COPY, "copy_generation_started")
        for product in self.repo.list_products(task_id):
            title = " · ".join(x for x in [product.product_name, product.category] if x)
            content = GeneratedContent(uuid4(), task_id, product.id, title, [f"类目：{product.category}"] if product.category else [], [], {"provider": "mock", "model": "deterministic-template-v1"}); self.repo.add_content(content)
        self.transition(task, TaskStatus.WAITING_COPY_REVIEW, "copy_generation_completed"); return self.workspace(task_id)
    @atomic
    def approve_copy(self, task_id: UUID, comment: str | None) -> dict[str, Any]:
        task = self.repo.get_task(task_id)
        if task.status != TaskStatus.WAITING_COPY_REVIEW: raise DomainError("INVALID_TASK_STATE", "当前状态不能审核文案", 409)
        if not self.repo.list_generated_content(task_id): raise DomainError("CONTENT_NOT_FOUND", "没有可审核文案", 409)
        approval = Approval(uuid4(), task_id, self.actor_id, "copy", "approved", comment); self.repo.add_approval(approval); self.transition(task, TaskStatus.APPROVED, "copy_approved"); return self.workspace(task_id)
    @atomic
    def export(self, task_id: UUID) -> TaskFile:
        task = self.repo.get_task(task_id)
        if task.status != TaskStatus.APPROVED: raise DomainError("INVALID_TASK_STATE", "当前状态不能导出", 409)
        item = TaskFile(uuid4(), task_id, self.storage.save("export", self.writer.export(self.workspace(task_id))), f"listing-{task_id}.xlsx", "export"); self.repo.add_file(item); self.transition(task, TaskStatus.EXPORTED, "export_created"); return item
    def download(self, task_id: UUID) -> tuple[TaskFile, bytes]:
        task = self.repo.get_task(task_id)
        if task.status != TaskStatus.EXPORTED: raise DomainError("INVALID_TASK_STATE", "当前状态不能下载", 409)
        items = [item for item in self.repo.list_task_files(task_id) if item.file_kind == "export"]
        if not items: raise DomainError("EXPORT_NOT_FOUND", "未找到导出文件", 404)
        item = max(items, key=lambda x: x.created_at); return item, self.storage.read(item.storage_path)
