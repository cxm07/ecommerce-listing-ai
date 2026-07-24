"""Integration coverage against the disposable Supabase database used by CI."""
from __future__ import annotations

import os
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from uuid import UUID, uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.config import Settings
from app.core import DomainError, WorkflowApplication
from app.storage import LocalStorageAdapter
from app.main import create_app
from app.persistence import PostgresRepositoryFactory
from app.workflow import TaskStatus


def _database_url() -> str:
    url = os.getenv("SUPABASE_DB_URL")
    if url:
        return url
    if os.getenv("CI"):
        pytest.fail("CI must provide SUPABASE_DB_URL for Postgres integration tests")
    pytest.skip("SUPABASE_DB_URL is supplied by the disposable CI Supabase instance")


@pytest.fixture
def postgres_factory() -> PostgresRepositoryFactory:
    actor = uuid4()
    factory = PostgresRepositoryFactory(_database_url(), actor)
    factory.open()
    try:
        with factory.pool.connection() as connection:
            connection.execute(
                "insert into auth.users(id,instance_id,aud,role,email,encrypted_password,email_confirmed_at,raw_app_meta_data,raw_user_meta_data,created_at,updated_at) values(%s,'00000000-0000-0000-0000-000000000000','authenticated','authenticated',%s,'test',now(),'{}','{}',now(),now())",
                (actor, f"{actor}@example.test"),
            )
            connection.execute("insert into public.profiles(id,display_name) values(%s,'integration')", (actor,))
            connection.commit()
        yield factory
    finally:
        factory.close()


@pytest.fixture
def workflow(postgres_factory: PostgresRepositoryFactory, tmp_path: Path) -> WorkflowApplication:
    return WorkflowApplication(postgres_factory, LocalStorageAdapter(tmp_path), str(postgres_factory.actor_id), 10 * 1024 * 1024)


@pytest.fixture
def sample_workbook() -> bytes:
    return (Path(__file__).resolve().parents[3] / "sample-data" / "sample-products.xlsx").read_bytes()


@pytest.mark.postgres_integration
def test_postgres_create_task_uses_single_transaction(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory) -> None:
    task = workflow.create_task("Persistent task", "test")
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).version == 1
        assert [event.action for event in repo.list_audit_logs(task.id)] == ["task_created"]


@pytest.mark.postgres_integration
def test_postgres_create_task_rolls_back_when_audit_fails(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, monkeypatch: pytest.MonkeyPatch) -> None:
    task_id = uuid4()
    monkeypatch.setattr("app.core.uuid4", lambda: task_id)
    monkeypatch.setattr(workflow, "audit", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("audit failure")))
    with pytest.raises(RuntimeError, match="audit failure"):
        workflow.create_task("rollback", "test")
    with postgres_factory.read_repository() as repo:
        with pytest.raises(DomainError, match="未找到任务"):
            repo.get_task(task_id)


@pytest.mark.postgres_integration
def test_postgres_upload_persists_file_status_version_and_audit(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes) -> None:
    task = workflow.create_task("Upload", "test")
    file = workflow.upload(task.id, "sample-products.xlsx", sample_workbook)
    with postgres_factory.read_repository() as repo:
        restored = repo.get_task(task.id)
        assert restored.status == TaskStatus.UPLOADED and restored.version == 2
        assert [item.id for item in repo.list_task_files(task.id)] == [file.id]
        assert {event.action for event in repo.list_audit_logs(task.id)} == {"task_created", "source_uploaded"}


@pytest.mark.postgres_integration
def test_postgres_upload_rolls_back_when_audit_fails(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes, monkeypatch: pytest.MonkeyPatch) -> None:
    task = workflow.create_task("Upload rollback", "test")
    monkeypatch.setattr(workflow, "audit", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("audit failure")))
    with pytest.raises(RuntimeError, match="audit failure"):
        workflow.upload(task.id, "sample-products.xlsx", sample_workbook)
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).status == TaskStatus.DRAFT
        assert repo.get_task(task.id).version == 1
        assert repo.list_task_files(task.id) == []
        assert [event.action for event in repo.list_audit_logs(task.id)] == ["task_created"]
    assert not list((workflow.storage.root / "tasks").rglob("source.xlsx"))


class _ConflictingWorkflow(WorkflowApplication):
    def advance_task(self, repo, task, new_status=None) -> int:
        with self.repository_factory.unit_of_work() as competing:
            assert competing.repository is not None
            competing.repository.advance_task(task.id, task.version)
            competing.commit()
        return super().advance_task(repo, task, new_status)


@pytest.mark.postgres_integration
def test_postgres_upload_rolls_back_on_version_conflict(postgres_factory: PostgresRepositoryFactory, tmp_path: Path, sample_workbook: bytes) -> None:
    workflow = _ConflictingWorkflow(postgres_factory, LocalStorageAdapter(tmp_path), str(postgres_factory.actor_id), 10 * 1024 * 1024)
    task = workflow.create_task("Conflict", "test")
    with pytest.raises(DomainError, match="其他请求"):
        workflow.upload(task.id, "sample-products.xlsx", sample_workbook)
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).status == TaskStatus.DRAFT
        assert repo.get_task(task.id).version == 2  # only the competing connection advanced it
        assert repo.list_task_files(task.id) == []
        assert [event.action for event in repo.list_audit_logs(task.id)] == ["task_created"]
    assert not list((tmp_path / "tasks").rglob("source.xlsx"))


@pytest.mark.postgres_integration
def test_postgres_parse_persists_complete_aggregate_and_rebuilds_application(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, tmp_path: Path, sample_workbook: bytes) -> None:
    task = workflow.create_task("Parse", "test")
    workflow.upload(task.id, "sample-products.xlsx", sample_workbook)
    summary = workflow.parse(task.id)
    assert summary == {"product_count": 1, "sku_count": 6, "issue_count": 5, "error_count": 2, "warning_count": 2, "info_count": 1}
    rebuilt = WorkflowApplication(postgres_factory, LocalStorageAdapter(tmp_path), str(postgres_factory.actor_id), 10 * 1024 * 1024)
    workspace = rebuilt.workspace(task.id)
    assert set(workspace) == {"task", "files", "products", "skus", "issues", "generated_content", "approvals", "audit_logs"}
    assert workspace["task"]["status"] == TaskStatus.WAITING_PRODUCT_REVIEW.value
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).version == 3
    assert len(workspace["files"]) == 1 and len(workspace["products"]) == 1
    assert len(workspace["skus"]) == 6 and len(workspace["issues"]) == 5
    assert {event["action"] for event in workspace["audit_logs"]} == {"task_created", "source_uploaded", "parsing_completed"}


@pytest.mark.postgres_integration
def test_postgres_parse_rolls_back_on_audit_failure(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes, monkeypatch: pytest.MonkeyPatch) -> None:
    task = workflow.create_task("Parse rollback", "test")
    workflow.upload(task.id, "sample-products.xlsx", sample_workbook)
    monkeypatch.setattr(workflow, "audit", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("audit failure")))
    with pytest.raises(RuntimeError, match="audit failure"):
        workflow.parse(task.id)
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).status == TaskStatus.UPLOADED
        assert repo.get_task(task.id).version == 2
        assert repo.list_products(task.id) == [] and repo.list_skus(task.id) == [] and repo.list_issues(task.id) == []
        assert {event.action for event in repo.list_audit_logs(task.id)} == {"task_created", "source_uploaded"}


@pytest.mark.postgres_integration
def test_postgres_parse_rolls_back_on_version_conflict(postgres_factory: PostgresRepositoryFactory, tmp_path: Path, sample_workbook: bytes) -> None:
    normal = WorkflowApplication(postgres_factory, LocalStorageAdapter(tmp_path), str(postgres_factory.actor_id), 10 * 1024 * 1024)
    task = normal.create_task("Parse conflict", "test")
    normal.upload(task.id, "sample-products.xlsx", sample_workbook)
    workflow = _ConflictingWorkflow(postgres_factory, LocalStorageAdapter(tmp_path), str(postgres_factory.actor_id), 10 * 1024 * 1024)
    with pytest.raises(DomainError, match="其他请求"):
        workflow.parse(task.id)
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).status == TaskStatus.UPLOADED
        assert repo.get_task(task.id).version == 3  # the competing connection, not parse, advanced it
        assert repo.list_products(task.id) == [] and repo.list_skus(task.id) == [] and repo.list_issues(task.id) == []


@pytest.mark.postgres_integration
def test_postgres_read_repository_supports_list_get_and_workspace(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory) -> None:
    task = workflow.create_task("Read", "test")
    with postgres_factory.read_repository() as repo:
        assert [item.id for item in repo.list_tasks() if item.id == task.id] == [task.id]
        assert repo.get_task(task.id).id == task.id
    assert workflow.workspace(task.id)["task"]["id"] == str(task.id)


@pytest.mark.postgres_integration
def test_postgres_api_task_responses_do_not_expose_version(workflow: WorkflowApplication) -> None:
    expected = {"id", "task_name", "category", "creator_id", "status", "created_at", "updated_at"}
    with TestClient(create_app(Settings(), workflow)) as client:
        created = client.post("/api/tasks", json={"task_name": "Public task", "category": "test"}).json()["data"]
        task_id = created["id"]
        values = [
            created,
            client.get("/api/tasks").json()["data"]["items"][0],
            client.get(f"/api/tasks/{task_id}").json()["data"],
            client.get(f"/api/tasks/{task_id}/workspace").json()["data"]["task"],
        ]
    for value in values:
        assert set(value) == expected
        assert "version" not in value


def _review_ready(workflow: WorkflowApplication, sample_workbook: bytes):
    task = workflow.create_task("Review", "test")
    workflow.upload(task.id, "sample-products.xlsx", sample_workbook)
    workflow.parse(task.id)
    return task, workflow.workspace(task.id)


@pytest.mark.postgres_integration
def test_postgres_patch_product_updates_issues_version_and_audit(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes) -> None:
    task, workspace = _review_ready(workflow, sample_workbook)
    product_id = workspace["products"][0]["id"]
    updated = workflow.patch_product(UUID(product_id), {"material": "combed cotton"})
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).version == 4
    assert updated["products"][0]["material"] == "combed cotton"
    assert any(item["action"] == "product_updated" for item in updated["audit_logs"])
    rebuilt = WorkflowApplication(postgres_factory, workflow.storage, workflow.actor_id, workflow.max_upload_bytes)
    assert rebuilt.workspace(task.id)["products"][0]["material"] == "combed cotton"


@pytest.mark.postgres_integration
def test_postgres_patch_product_rolls_back_when_audit_fails(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes, monkeypatch: pytest.MonkeyPatch) -> None:
    task, workspace = _review_ready(workflow, sample_workbook)
    product_id = workspace["products"][0]["id"]
    monkeypatch.setattr(workflow, "audit", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("audit failure")))
    with pytest.raises(RuntimeError, match="audit failure"):
        workflow.patch_product(UUID(product_id), {"material": "new material"})
    with postgres_factory.read_repository() as repo:
        assert repo.get_product(product_id).material != "new material"
        assert repo.get_task(task.id).version == 3
        assert not any(item.action == "product_updated" for item in repo.list_audit_logs(task.id))


@pytest.mark.postgres_integration
def test_postgres_patch_product_rolls_back_on_version_conflict(postgres_factory: PostgresRepositoryFactory, tmp_path: Path, sample_workbook: bytes) -> None:
    normal = WorkflowApplication(postgres_factory, LocalStorageAdapter(tmp_path), str(postgres_factory.actor_id), 10 * 1024 * 1024)
    task, workspace = _review_ready(normal, sample_workbook)
    conflicting = _ConflictingWorkflow(postgres_factory, normal.storage, normal.actor_id, normal.max_upload_bytes)
    with pytest.raises(DomainError, match="其他请求"):
        conflicting.patch_product(UUID(workspace["products"][0]["id"]), {"material": "new material"})
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).version == 4
        assert repo.get_product(UUID(workspace["products"][0]["id"])).material != "new material"
        assert not any(item.action == "product_updated" for item in repo.list_audit_logs(task.id))


@pytest.mark.postgres_integration
def test_postgres_patch_sku_updates_issues_and_preserves_decimal(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes) -> None:
    task, workspace = _review_ready(workflow, sample_workbook)
    duplicate = next(item for item in workspace["issues"] if item["code"] == "DUPLICATE_SKU")
    invalid_price = next(item for item in workspace["issues"] if item["code"] == "INVALID_PRICE")
    updated = workflow.patch_sku(UUID(duplicate["sku_id"]), {"sku_code": "TSHIRT-WHITE-XL"})
    updated = workflow.patch_sku(UUID(invalid_price["sku_id"]), {"price": Decimal("79.90"), "color": "white"})
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).version == 5
    corrected = next(item for item in updated["skus"] if item["id"] == invalid_price["sku_id"])
    assert corrected["price"] == 79.9
    states = {item["code"]: item["resolved"] for item in updated["issues"]}
    assert states["DUPLICATE_SKU"] and states["INVALID_PRICE"]
    assert any(item["action"] == "sku_updated" for item in updated["audit_logs"])
    assert updated["task"]["id"] == str(task.id)


@pytest.mark.postgres_integration
def test_postgres_patch_sku_rolls_back_on_version_conflict(postgres_factory: PostgresRepositoryFactory, tmp_path: Path, sample_workbook: bytes) -> None:
    normal = WorkflowApplication(postgres_factory, LocalStorageAdapter(tmp_path), str(postgres_factory.actor_id), 10 * 1024 * 1024)
    task, workspace = _review_ready(normal, sample_workbook)
    target = workspace["skus"][0]
    conflicting = _ConflictingWorkflow(postgres_factory, normal.storage, normal.actor_id, normal.max_upload_bytes)
    with pytest.raises(DomainError, match="其他请求"):
        conflicting.patch_sku(UUID(target["id"]), {"color": "blue"})
    with postgres_factory.read_repository() as repo:
        assert repo.get_sku(target["id"]).color == target["color"]
        assert repo.get_task(task.id).version == 4
        assert not any(item.action == "sku_updated" for item in repo.list_audit_logs(task.id))


@pytest.mark.postgres_integration
def test_postgres_patch_sku_rolls_back_when_audit_fails(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes, monkeypatch: pytest.MonkeyPatch) -> None:
    task, workspace = _review_ready(workflow, sample_workbook)
    target = workspace["skus"][0]
    monkeypatch.setattr(workflow, "audit", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("audit failure")))
    with pytest.raises(RuntimeError, match="audit failure"):
        workflow.patch_sku(UUID(target["id"]), {"color": "blue"})
    with postgres_factory.read_repository() as repo:
        assert repo.get_sku(UUID(target["id"])).color == target["color"]
        assert repo.get_task(task.id).version == 3
        assert not any(item.action == "sku_updated" for item in repo.list_audit_logs(task.id))


@pytest.mark.postgres_integration
def test_postgres_approve_products_is_blocked_by_unresolved_errors(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes) -> None:
    task, _ = _review_ready(workflow, sample_workbook)
    with pytest.raises(DomainError, match="错误级问题") as error:
        workflow.approve_products(task.id, "blocked")
    assert error.value.code == "UNRESOLVED_ERROR_ISSUES"
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).status == TaskStatus.WAITING_PRODUCT_REVIEW
        assert repo.get_task(task.id).version == 3
        assert repo.list_approvals(task.id) == []


@pytest.mark.postgres_integration
def test_postgres_approve_products_persists_approval_status_version_and_audit(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes) -> None:
    task, workspace = _review_ready(workflow, sample_workbook)
    duplicate = next(item for item in workspace["issues"] if item["code"] == "DUPLICATE_SKU")
    invalid_price = next(item for item in workspace["issues"] if item["code"] == "INVALID_PRICE")
    workflow.patch_sku(UUID(duplicate["sku_id"]), {"sku_code": "TSHIRT-WHITE-XL"})
    workflow.patch_sku(UUID(invalid_price["sku_id"]), {"price": Decimal("79.90")})
    approved = workflow.approve_products(task.id, "approved by test")
    assert approved["task"]["status"] == TaskStatus.PRODUCT_APPROVED.value
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).version == 6
    assert approved["approvals"][0]["approval_type"] == "product"
    assert approved["approvals"][0]["comment"] == "approved by test"
    rebuilt = WorkflowApplication(postgres_factory, workflow.storage, workflow.actor_id, workflow.max_upload_bytes)
    assert rebuilt.workspace(task.id)["task"]["status"] == TaskStatus.PRODUCT_APPROVED.value


@pytest.mark.postgres_integration
def test_postgres_approve_products_rolls_back_when_audit_fails(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes, monkeypatch: pytest.MonkeyPatch) -> None:
    task, workspace = _review_ready(workflow, sample_workbook)
    duplicate = next(item for item in workspace["issues"] if item["code"] == "DUPLICATE_SKU")
    invalid_price = next(item for item in workspace["issues"] if item["code"] == "INVALID_PRICE")
    workflow.patch_sku(UUID(duplicate["sku_id"]), {"sku_code": "TSHIRT-WHITE-XL"})
    workflow.patch_sku(UUID(invalid_price["sku_id"]), {"price": Decimal("79.90")})
    monkeypatch.setattr(workflow, "audit", lambda *args, **kwargs: (_ for _ in ()).throw(RuntimeError("audit failure")))
    with pytest.raises(RuntimeError, match="audit failure"):
        workflow.approve_products(task.id, "no commit")
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).status == TaskStatus.WAITING_PRODUCT_REVIEW
        assert repo.get_task(task.id).version == 5
        assert repo.list_approvals(task.id) == []


@pytest.mark.postgres_integration
def test_postgres_approve_products_rolls_back_on_version_conflict(postgres_factory: PostgresRepositoryFactory, tmp_path: Path, sample_workbook: bytes) -> None:
    normal = WorkflowApplication(postgres_factory, LocalStorageAdapter(tmp_path), str(postgres_factory.actor_id), 10 * 1024 * 1024)
    task, workspace = _review_ready(normal, sample_workbook)
    duplicate = next(item for item in workspace["issues"] if item["code"] == "DUPLICATE_SKU")
    invalid_price = next(item for item in workspace["issues"] if item["code"] == "INVALID_PRICE")
    normal.patch_sku(UUID(duplicate["sku_id"]), {"sku_code": "TSHIRT-WHITE-XL"})
    normal.patch_sku(UUID(invalid_price["sku_id"]), {"price": Decimal("79.90")})
    conflicting = _ConflictingWorkflow(postgres_factory, normal.storage, normal.actor_id, normal.max_upload_bytes)
    with pytest.raises(DomainError, match="其他请求"):
        conflicting.approve_products(task.id, "conflict")
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).status == TaskStatus.WAITING_PRODUCT_REVIEW
        assert repo.get_task(task.id).version == 6
        assert repo.list_approvals(task.id) == []


@pytest.mark.postgres_integration
def test_complete_v1_workflow_with_postgres(workflow: WorkflowApplication, postgres_factory: PostgresRepositoryFactory, sample_workbook: bytes) -> None:
    task, workspace = _review_ready(workflow, sample_workbook)
    duplicate = next(item for item in workspace["issues"] if item["code"] == "DUPLICATE_SKU")
    invalid_price = next(item for item in workspace["issues"] if item["code"] == "INVALID_PRICE")
    workflow.patch_sku(UUID(duplicate["sku_id"]), {"sku_code": "TSHIRT-WHITE-XL"})
    workflow.patch_sku(UUID(invalid_price["sku_id"]), {"price": Decimal("79.90")})
    workflow.approve_products(task.id, "products approved")
    generated = workflow.generate_copy(task.id)
    assert len(generated["generated_content"]) == len(generated["products"])
    workflow.approve_copy(task.id, "copy approved")
    exported = workflow.export(task.id)
    item, payload = workflow.download(task.id)
    assert item.id == exported.id
    assert set(load_workbook(BytesIO(payload)).sheetnames) == {"products", "skus", "listing-copy", "issues", "audit-summary"}
    rebuilt = WorkflowApplication(postgres_factory, workflow.storage, workflow.actor_id, workflow.max_upload_bytes)
    restored = rebuilt.workspace(task.id)
    assert restored["task"]["status"] == TaskStatus.EXPORTED.value
    with postgres_factory.read_repository() as repo:
        assert repo.get_task(task.id).version == 9
    assert len(restored["generated_content"]) == len(restored["products"])
    assert {approval["approval_type"] for approval in restored["approvals"]} == {"product", "copy"}
