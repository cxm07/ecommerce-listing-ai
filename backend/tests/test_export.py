from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from conftest import assert_envelope


def test_export_and_download_xlsx_after_full_workflow(client: TestClient, sample_workbook: bytes) -> None:
    task_id = client.post("/api/tasks", json={"task_name": "夏季上新", "category": "服饰"}).json()["data"]["id"]
    client.post(f"/api/tasks/{task_id}/files", files={"file": ("sample-products.xlsx", sample_workbook)})
    issues = client.post(f"/api/tasks/{task_id}/parse").json()["issues"]
    duplicate = next(issue for issue in issues if issue["code"] == "DUPLICATE_SKU")
    invalid_price = next(issue for issue in issues if issue["code"] == "INVALID_PRICE")
    client.patch(f"/api/skus/{duplicate['sku_id']}", json={"sku_code": "TSHIRT-WHITE-XL"})
    client.patch(f"/api/skus/{invalid_price['sku_id']}", json={"price": 79.90})
    client.post(f"/api/tasks/{task_id}/approve-products", json={"decision": "approved"})
    client.post(f"/api/tasks/{task_id}/generate-copy")
    client.post(f"/api/tasks/{task_id}/approve-copy", json={"decision": "approved"})

    exported = client.post(f"/api/tasks/{task_id}/export")
    assert exported.status_code == 200
    assert_envelope(exported.json())
    assert exported.json()["data"]["file_id"]
    downloaded = client.get(f"/api/tasks/{task_id}/download")
    assert downloaded.status_code == 200
    assert downloaded.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "attachment;" in downloaded.headers["content-disposition"]
    assert load_workbook(BytesIO(downloaded.content)).sheetnames == ["products", "skus", "listing-copy", "issues", "audit-summary"]


def test_download_failure_is_json_envelope(client: TestClient) -> None:
    task_id = client.post("/api/tasks", json={"task_name": "夏季上新", "category": "服饰"}).json()["data"]["id"]
    response = client.get(f"/api/tasks/{task_id}/download")
    assert response.status_code == 409
    assert response.headers["content-type"].startswith("application/json")
    assert_envelope(response.json())
    assert response.json()["error"]["code"] == "INVALID_TASK_STATE"
